from flask import Flask, request,jsonify
from flask import send_file
from openpyxl import Workbook, load_workbook
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "parking_data.xlsx"

# Create Excel file if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parking Data"
    ws.append([
        "Date",
        "Vehicle Type",
        "Personnel",
        "Passengers",
        "Vehicle Number",
        "Entry Time",
        "Exit Time",
        "Extra Hours",
        "Total Hours",
        "Total Cost",
        "Remarks"
    ])
    wb.save(EXCEL_FILE)


@app.route("/submit", methods=["POST"])
def submit_data():
    data = request.json
    print("Recieved Data:",data)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        data.get("date"),
        data.get("vehicle"),
        data.get("personnel"),
        data.get("passengers"),
        data.get("vehicleNo"),
        data.get("entryTime"),
        data.get("exitTime"),
        data.get("extra"),
        data.get("totalHours"),
        data.get("totalCost"),
        data.get("remarks")
    ])

    wb.save(EXCEL_FILE)

    return jsonify({"success": True, "message": "Data saved to Excel"})

@app.route("/excel", methods=["GET"])
def open_excel():
    return send_file(EXCEL_FILE, as_attachment=False)


@app.route("/report", methods=["GET"])
def get_report():
    filter_date = request.args.get("date", "").strip()
    filter_vehicle_no = request.args.get("vehicleNo", "").strip().upper()

    if not os.path.exists(EXCEL_FILE):
        return jsonify({"success": False, "message": "No data file found"}), 404

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Each row: Date, Vehicle Type, Personnel, Passengers, Vehicle Number,
        #           Entry Time, Exit Time, Extra Hours, Total Hours, Total Cost, Remarks
        if not any(row):
            continue

        row_date = str(row[0]).strip() if row[0] else ""
        row_vehicle_no = str(row[4]).strip().upper() if row[4] else ""

        # Apply date filter
        if filter_date and row_date != filter_date:
            continue

        # Apply vehicle number filter (partial match, case-insensitive)
        if filter_vehicle_no and filter_vehicle_no not in row_vehicle_no:
            continue

        rows.append({
            "date": row[0],
            "vehicle": row[1],
            "personnel": row[2],
            "passengers": row[3],
            "vehicleNo": row[4],
            "entryTime": row[5],
            "exitTime": row[6],
            "extraHours": row[7],
            "totalHours": row[8],
            "totalCost": row[9],
            "remarks": row[10]
        })

    # Calculate total collection for the filtered results
    total_collection = 0
    for r in rows:
        try:
            total_collection += float(r["totalCost"]) if r["totalCost"] else 0
        except (ValueError, TypeError):
            pass

    return jsonify({
        "success": True,
        "data": rows,
        "totalCollection": round(total_collection, 2),
        "count": len(rows)
    })


if __name__ == "__main__":
    app.run(debug=True)
